from openpyxl import Workbook
from openpyxl.chart import LineChart, Reference
from openpyxl.styles import Alignment

def plot():
    # 读取文件
    with open('iostat_all.log', 'r') as f:
        lines = f.readlines()

    # 创建一个Workbook对象
    wb = Workbook()
    ws = wb.active

    # 填充数据
    first_columns = set()
    for line in lines:
        columns = line.split()
        if len(columns) <= 0:
            continue
        first_column = columns[0]
        if first_column.startswith('sd'):
            first_columns.add(first_column)

    columns_sort = list(first_columns)
    columns_sort.sort()

    column_num = 1
    for sd in columns_sort:
        column_num += 1
        row_counter = 2
        for line in lines:
            if sd in line:
                columns = line.split()
                third_column = columns[2]
                ws.cell(row=1, column=column_num).value = sd
                cell = ws.cell(row=row_counter, column=column_num)
                cell.value = float(third_column)
                row_counter += 1

    ws['A1'] = '收集数据次数'
    max_row = ws.max_row

    for i in range(2, max_row + 1):
        ws.cell(row=i, column=1).value = i - 1

    cell_range = ws[ws.dimensions]
    for row in cell_range:
        for cell in row:
            cell.alignment = Alignment(horizontal='left')

    # 设置数据范围和 x 轴范围
    data_range = Reference(ws, min_col=2, min_row=1, max_col=column_num, max_row=row_counter - 1)
    # 创建Reference对象
    x_axis_range = Reference(ws, min_col=1, min_row=2, max_row=row_counter - 1)

    # 创建折线图对象
    chart = LineChart()
    chart.add_data(data_range, titles_from_data=True)
    # 使用手动创建的Reference对象
    chart.set_categories(x_axis_range)
    chart.title = "Hot_Plug IO Performance"
    chart.x_axis.title = 'Iostat Performance Collect 5s/次'
    chart.y_axis.title = 'Iostat Performance w/s'

    # 设置图表样式，只显示坐标轴，不显示网格线
    chart.y_axis.majorGridlines = None
    chart.x_axis.majorGridlines = None

    # 将图表添加到工作表
    ws.add_chart(chart, "M7")

    # 保存Excel文件
    wb.save("iostat_chart.xlsx")


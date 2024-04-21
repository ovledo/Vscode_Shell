import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill
import sys

filename = sys.argv[1]

df = pd.read_excel(filename)
df.to_excel('add_contrast.xlsx',index=False)

wb = openpyxl.load_workbook('add_contrast.xlsx')
ws = wb.active

for cell in ws[1]:
    cell.font = None
    cell.border = None

contrast_list = []
for column in df.columns[2:]:
    # 获取当前列的所有值
    values = df[column].values
    column_Max = values.max()
    column_Min = values.min()

    con = column_Min / column_Max
    contrast_list.append(con)

max_row_num = ws.max_row
ws._current_row = max_row_num

ws.append([None, None] + contrast_list)

# 遍历所有单元格
for row in ws.iter_rows():
    for cell in row:
        # 检查单元格的值是否小于0.95
        if isinstance(cell.value, (int, float)) and cell.value < 0.95:
            # 将单元格标记为红色
            cell.fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")


wb.save('add_contrast.xlsx')

df = pd.read_excel('add_contrast.xlsx')
print(df)
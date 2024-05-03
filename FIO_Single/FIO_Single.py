import os
import pandas as pd
import re
import openpyxl
from openpyxl.chart import LineChart, Reference
from openpyxl import load_workbook

def process_sd_csv_files(directory):
    sd_csv_files = []
    for filename in os.listdir(directory):
        if filename.endswith(".csv") and filename.startswith("sd"):
            filepath = os.path.join(directory, filename)
            sd_csv_files.append(filepath)
    return sd_csv_files

directory = os.getcwd()
sd_csv_files = process_sd_csv_files(directory)

def Fill(filename):
    wb = openpyxl.Workbook()
    default_sheet = wb["Sheet"]
    wb.remove(default_sheet)

    Test_Mode = ["read", "write", "randread", "randwrite"]
    for new_sheet in Test_Mode:
        df = pd.read_csv(filename)
        df = df[df["Test-Mode"] == new_sheet]

        # 确定工作表名称
        if new_sheet in ["read", "write"]:
            sheet_name = "Seq_" + new_sheet
        else:
            sheet_name = "Random_" + new_sheet.replace("rand", "")

        # 创建一个新的工作表来存储数据
        if sheet_name in wb.sheetnames:
            wb.remove(wb[sheet_name])
        ws = wb.create_sheet(title=sheet_name)

        # 写入表头
        if new_sheet in ["randread", "randwrite"]:
            column_names = ["4k", "8k", "16k", "32k", "64k"]
            data_column = "IOPS"  # 指定数据列为 "IOPS"
        else:
            column_names = ["4k", "8k", "16k", "32k", "64k", "128k", "256k", "512k", "1m", "2m", "4m", "8m", "16m"]
            data_column = "Bandwidth"  # 指定数据列为 "Bandwidth"

        ws.append([""] + column_names)  # 添加空值作为第一列的占位符

        # 存储数据的列表
        data_list = []

        # 按照每个 data 的值进行筛选数据，并生成对应的 New_data
        for data_str in ["1", "2", "4", "8", "16", "32"]:
            data = int(data_str)
            # 存储当前 data 对应的 New_data
            current_data = ["QD=" + data_str]
            for column_list in column_names:
                # 筛选满足 Blocksize 列为 column_list 且 Queue-Depth 列为 "QD=data" 的数据
                filtered_data = df[(df["Blocksize"].str.contains('^' + re.escape(column_list) + '$')) & (
                        df["Queue-Depth"] == data)]
                New_data = filtered_data[data_column].tolist()  # 使用指定的数据列
                # 删除 "IOPS" 或 "MiB" 后缀并转换为浮点数
                New_data = [float(value.replace("IOPS", "").replace("MiB", "")) for value in New_data]
                # 如果 New_data 为空，则插入一个空字符串
                if not New_data:
                    current_data.append("")
                else:
                    # 否则插入 New_data 中的第一个元素
                    current_data.append(New_data[0])
            # 将当前 data 对应的 New_data 添加到数据列表中
            ws.append(current_data)

    # 保存Excel文件
    wb.save('FIO_Single.xlsx')

for csv in sd_csv_files:
    Fill(csv)


def draw_line_chart(filename):
    # 加载现有的Excel文件
    wb = load_workbook(filename)

    # 遍历每个工作表
    for ws in wb.worksheets:

        chart = LineChart()
        # 设置图表标题
        chart.title = ws.title
        cur_max_row = ws.max_row
        cur_max_column = ws.max_column
        datas = Reference(ws, min_row=2, max_row=cur_max_row, min_col=1, max_col=cur_max_column)
        line_data = Reference(ws, min_row=1, max_row=1, min_col=2, max_col=14)
        chart.add_data(datas, from_rows=True, titles_from_data=True)
        chart.set_categories(line_data)
        # 将图表添加到工作表中
        ws.add_chart(chart, "M4")

    # 保存Excel文件
    wb.save(filename)


# 测试函数
draw_line_chart('FIO_Single.xlsx')